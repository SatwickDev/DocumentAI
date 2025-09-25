# ---------------- Imports ----------------
import os
import fitz  # PyMuPDF
import pytesseract
import cv2
import numpy as np
from PIL import Image
import json
import gradio as gr
import shutil
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Dict, List, Tuple, Optional
import io
from docx import Document
from openpyxl import load_workbook
import logging
import gc
import uuid
import argparse

# ---------------- Logging Setup ----------------
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("document_classification.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ---------------- Configurable Paths via Environment Variables ----------------
# Try multiple common Tesseract installation paths
tesseract_paths = [
    r"C:\Program Files\Tesseract-OCR\tesseract.exe",
    r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
    os.environ.get('TESSERACT_CMD')
]

for path in tesseract_paths:
    if path and os.path.exists(path):
        TESSERACT_CMD = path
        break
else:
    TESSERACT_CMD = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
    logger.warning(f"Tesseract not found in common locations. Using default path: {TESSERACT_CMD}")

pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD

OUTPUT_BASE_PATH = os.environ.get('OUTPUT_BASE_PATH', r"C:\DocumentClassificationOutput")

CONFIG_PATH_ENV = os.environ.get('CONFIG_PATH')

# ---------------- Load Classification Configuration from JSON ----------------
def load_classification_config(config_path: Optional[str] = None) -> Dict:
    """
    Load classification configuration with validation for categories, keywords, and splitting rules.
    """
    if config_path is None:
        config_path = CONFIG_PATH_ENV or os.path.join(os.path.dirname(os.path.abspath(__file__)), "classification_config.json")
    
    logger.info(f"Looking for config at: {config_path}")
    logger.debug(f"Script directory: {os.path.dirname(os.path.abspath(__file__))}")
    logger.debug(f"Current working directory: {os.getcwd()}")
    logger.debug(f"Config file exists: {os.path.exists(config_path)}")
    
    try:
        with open(config_path, "r", encoding="utf-8") as f:
            config = json.load(f)
        
        logger.info(f"Successfully loaded {len(config['categories'])} document categories from {config_path}")
        
        logger.debug("Loaded categories and sample keywords:")
        for category, data in config["categories"].items():
            keywords = data.get("keywords", [])
            keyword_count = len(keywords)
            sample = keywords[:3]
            logger.debug(f"   • {category}: {keyword_count} keywords - {sample}")
        
        compiled_config = {}
        always_separate_categories = []
        category_page_limits = {}
        
        for category, data in config["categories"].items():
            if data.get("always_separate", False):
                always_separate_categories.append(category)
            category_page_limits[category] = data.get("max_pages_per_pdf", config.get("default_pages_per_pdf", 1))
            
            keywords = data.get("keywords", [])
            compiled_keywords = []
            for item in keywords:
                if isinstance(item, dict) and "keyword" in item:
                    kw = item["keyword"].lower()
                    weight = item.get("weight", 1.0)
                    compiled_keywords.append({"keyword": kw, "normalized": ''.join(c for c in kw if c.isalnum()), "weight": weight})
                elif isinstance(item, str):
                    kw = item.lower()
                    compiled_keywords.append({"keyword": kw, "normalized": ''.join(c for c in kw if c.isalnum()), "weight": 1.0})
            compiled_config[category] = sorted(compiled_keywords, key=lambda x: len(x["keyword"]), reverse=True)
        
        return {
            "categories": compiled_config,
            "always_separate_categories": always_separate_categories,
            "category_page_limits": category_page_limits,
            "default_pages_per_pdf": config.get("default_pages_per_pdf", 1),
            "tesseract_configs": config.get("tesseract_configs", [r'--oem 3 --psm 6 -c tessedit_char_whitelist=0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz.,:-/()']),
            "min_text_length": config.get("min_text_length", 5)
        }
        
    except FileNotFoundError:
        logger.error(f"Configuration file {config_path} not found!")
        logger.debug("Files in script directory:")
        script_dir = os.path.dirname(config_path)
        if os.path.exists(script_dir):
            for file in os.listdir(script_dir):
                if file.endswith(('.json', '.py')):
                    logger.debug(f"   • {file}")
        raise FileNotFoundError("Configuration file is not present. Hence classification is aborted.")
    
    except json.JSONDecodeError as e:
        logger.error(f"Invalid JSON in {config_path}: {e}")
        logger.debug("Please check your JSON syntax - common issues:")
        logger.debug("   • Missing commas between items")
        logger.debug("   • Unmatched brackets or quotes")
        logger.debug("   • Trailing commas")
        return {
            "categories": {},
            "always_separate_categories": [],
            "category_page_limits": {},
            "default_pages_per_pdf": 1,
            "tesseract_configs": [r'--oem 3 --psm 6 -c tessedit_char_whitelist=0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz.,:-/()'],
            "min_text_length": 5
        }

# Load configuration at startup
try:
    config = load_classification_config()
    classification_keywords = config["categories"]
    always_separate_categories = config["always_separate_categories"]
    category_page_limits = config["category_page_limits"]
    default_pages_per_pdf = config["default_pages_per_pdf"]
    tesseract_configs = config["tesseract_configs"]
    min_text_length = config["min_text_length"]
except Exception as e:
    logger.critical(f"Failed to load configuration: {str(e)}")
    raise

# ---------------- Helper Functions ----------------

def get_max_pages_for_category(category: str) -> int:
    """
    Get maximum pages allowed per PDF for a given category.
    """
    return category_page_limits.get(category, default_pages_per_pdf)

class PageProcessor:
    """Optimized page processor with memory-based operations and configurable OCR."""
    
    def __init__(self):
        self.tesseract_configs = tesseract_configs
        
    def classify_page_type_fast(self, page) -> Tuple[str, str]:
        """
        Ultra-fast page classification based on text content and dimensions.
        """
        try:
            text = page.get_text()
            if len(text.strip()) > 50:
                return "text_based", text
            rect = page.rect
            area = rect.width * rect.height
            return ("high_quality_scan" if area > 800000 else "low_quality_scan", text)
        except Exception as e:
            logger.error(f"Error classifying page type: {e}")
            return "unknown", ""

    def process_image_in_memory(self, pix) -> str:
        """
        Process image entirely in memory with robust preprocessing and multiple OCR configs.
        """
        try:
            if pix.n == 1:
                mode = "L"
            elif pix.n == 3:
                mode = "RGB"
            elif pix.n == 4:
                mode = "RGBA"
            else:
                mode = "RGB"
            
            pil_image = Image.frombytes(mode, [pix.width, pix.height], pix.samples)
            
            if pil_image.mode != 'RGB':
                pil_image = pil_image.convert('RGB')
            
            img_array = np.array(pil_image)
            
            if len(img_array.shape) == 3:
                gray = cv2.cvtColor(img_array, cv2.COLOR_RGB2GRAY)
            else:
                gray = img_array
            
            # Enhanced preprocessing
            enhanced = cv2.convertScaleAbs(gray, alpha=1.2, beta=20)
            cleaned = cv2.medianBlur(enhanced, 5)
            thresh = cv2.adaptiveThreshold(cleaned, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2)
            sharpened = cv2.filter2D(thresh, -1, np.array([[0, -1, 0], [-1, 5, -1], [0, -1, 0]]))
            
            processed_pil = Image.fromarray(sharpened)
            
            # Try multiple Tesseract configs
            for config in self.tesseract_configs:
                try:
                    text = pytesseract.image_to_string(processed_pil, config=config)
                    if text.strip():
                        return text
                except Exception as e:
                    logger.warning(f"OCR failed with config {config}: {e}")
                    continue
            
            # Fallback methods
            try:
                pil_image = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                text = pytesseract.image_to_string(pil_image, config=self.tesseract_configs[0])
                if text.strip():
                    return text
            except Exception as e:
                logger.warning(f"Fallback 1 failed: {e}")
                try:
                    pil_image = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                    return pytesseract.image_to_string(pil_image)
                except Exception as e2:
                    logger.warning(f"Fallback 2 failed: {e2}")
                    return ""
            
            return ""
            
        except Exception as e:
            logger.error(f"Memory OCR failed: {e}")
            return ""
        finally:
            gc.collect()

    def extract_text_ultra_fast(self, page, page_number: int, session_output_path: str) -> str:
        """
        Ultra-optimized text extraction with zero file I/O and debug logging.
        """
        page_type, existing_text = self.classify_page_type_fast(page)
        
        if page_type == "text_based":
            return existing_text
        
        try:
            matrix = fitz.Matrix(2.0, 2.0) if page_type == "low_quality_scan" else fitz.Matrix(1.5, 1.5)
            pix = page.get_pixmap(matrix=matrix)
            
            text = self.process_image_in_memory(pix)
            
            return text
        except Exception as e:
            logger.error(f"Error extracting text from page {page_number}: {e}")
            return ""
        finally:
            pix = None
            gc.collect()

def normalize_text_for_matching(text: str) -> str:
    """
    Aggressive text normalization to handle OCR variations.
    """
    if not text:
        return ""
    
    normalized = ''.join(char.lower() for char in text if char.isalnum())
    return normalized

def fuzzy_keyword_match(text_normalized: str, keyword_normalized: str, threshold: float = 0.75) -> bool:
    """
    Stricter fuzzy matching to reduce false positives.
    """
    if not keyword_normalized or not text_normalized:
        return False
    
    if keyword_normalized in text_normalized:
        return True
    
    if len(keyword_normalized) < 8:
        return False
    
    keyword_chars = list(keyword_normalized)
    text_chars = list(text_normalized)
    
    matches = 0
    text_idx = 0
    consecutive_matches = 0
    max_consecutive = 0
    
    for kw_char in keyword_chars:
        found = False
        while text_idx < len(text_chars):
            if text_chars[text_idx] == kw_char:
                matches += 1
                text_idx += 1
                found = True
                consecutive_matches += 1
                max_consecutive = max(max_consecutive, consecutive_matches)
                break
            text_idx += 1
        
        if not found:
            consecutive_matches = 0
    
    similarity = matches / len(keyword_chars)
    consecutive_ratio = max_consecutive / len(keyword_chars)
    
    return similarity >= threshold and consecutive_ratio >= 0.4

def classify_text_ultra_fast(text: str, page_num: str) -> str:
    """
    Robust and precise classification with weighted keywords and tiebreaker.
    """
    if not text or len(text.strip()) < min_text_length:
        logger.debug(f"Text too short for classification (page {page_num}): '{text[:50]}...'")
        return "Unclassified"
    
    logger.debug(f"Raw text for page {page_num} (first 200 chars): '{text[:200]}...'")
    text_clean = ' '.join(text.lower().split())
    logger.debug(f"Original text (first 100 chars): '{text_clean[:100]}...'")
    
    text_normalized = normalize_text_for_matching(text)
    logger.debug(f"Normalized text (first 100 chars): '{text_normalized[:100]}...'")
    logger.debug(f"Using {len(classification_keywords)} categories from JSON config")
    
    category_scores = {}
    matches_found = []
    
    for category, keywords in classification_keywords.items():
        if not keywords:
            continue
            
        category_score = 0
        exact_matches = 0
        category_matches = []
        
        for kw_item in keywords:
            keyword = kw_item["keyword"]
            keyword_normalized = kw_item["normalized"]
            weight = kw_item["weight"]
            
            match_found = False
            match_type = ""
            match_score = 0
            
            if keyword_normalized in text_normalized:
                match_found = True
                match_type = "exact"
                exact_matches += 1
                match_score = len(keyword_normalized) * 4 * weight
            
            elif len(keyword_normalized) >= 8 and fuzzy_keyword_match(text_normalized, keyword_normalized, threshold=0.75):
                match_found = True
                match_type = "fuzzy"
                match_score = len(keyword_normalized) * 1.5 * weight
            
            elif len(keyword_normalized) >= 6:
                keyword_parts = keyword.split()
                if len(keyword_parts) >= 2:
                    parts_found = sum(1 for part in keyword_parts 
                                    if len(normalize_text_for_matching(part)) >= 3 and 
                                       normalize_text_for_matching(part) in text_normalized)
                    if parts_found >= len(keyword_parts) * 0.8:
                        match_found = True
                        match_type = "partial"
                        match_score = len(keyword_normalized) * 1.0 * weight
            
            if match_found:
                category_score += match_score
                category_matches.append(f"{keyword}({match_type},w={weight})")
        
        if category_score > 0:
            if exact_matches > 0:
                category_score *= (1 + exact_matches * 0.5)
                
            matches_found.append(f"{category}: {category_matches} = {category_score:.1f} (exact: {exact_matches})")
            category_scores[category] = {"score": category_score, "exact_matches": exact_matches}
    
    if not category_scores:
        logger.debug(f"No keyword matches found for page {page_num}")
        logger.debug(f"Sample normalized keywords to check:")
        for cat, keywords in list(classification_keywords.items())[:2]:
            if keywords:
                sample_kw = keywords[0]["keyword"]
                logger.debug(f"     {cat}: '{sample_kw}' → '{keywords[0]['normalized']}'")
        return "Unclassified"
    
    best_category = max(category_scores, key=lambda c: (category_scores[c]["score"], category_scores[c]["exact_matches"]))
    best_score = category_scores[best_category]["score"]
    
    logger.debug(f"All matches for page {page_num}: {matches_found}")
    logger.info(f"Best match for page {page_num}: {best_category} (score: {best_score:.1f})")
    
    return best_category

def process_page_batch(pages_data: List[Tuple], session_output_path: str) -> List[Dict]:
    """
    Process a batch of pages with detailed debugging.
    """
    processor = PageProcessor()
    results = []
    
    for page, page_num in pages_data:
        try:
            logger.info(f"Processing page {page_num}...")
            
            text = processor.extract_text_ultra_fast(page, page_num, session_output_path)
            logger.debug(f"Extracted {len(text)} characters")
            
            if text.strip():
                logger.debug(f"First 150 chars: '{text[:150]}...'")
            else:
                logger.warning(f"No text extracted from page {page_num}")
            
            category = classify_text_ultra_fast(text, page_num)
            quality = "Good" if text.strip() and len(text.split()) >= 3 else "Poor Quality"
            
            logger.info(f"Page {page_num}: {category} ({quality})")
            
            results.append({
                "page_num": page_num,
                "text": text,
                "category": category,
                "quality": quality,
                "text_length": len(text)
            })
            
        except Exception as e:
            logger.error(f"Error processing page {page_num}: {e}")
            with open(os.path.join(session_output_path, f"error_page_{page_num}.log"), "w", encoding="utf-8") as f:
                f.write(f"Error: {str(e)}\n")
            results.append({
                "page_num": page_num,
                "text": "",
                "category": "Unclassified",
                "quality": "Error",
                "text_length": 0
            })
        finally:
            gc.collect()
    
    return results

def create_pdfs_batch_optimized(doc, results: List[Dict], output_path: str, session_id: str) -> List[Dict]:
    """
    Ultra-fast PDF creation with batch operations and unique filenames.
    """
    category_groups = {}
    classification_log = []
    
    for result in results:
        category = result["category"]
        page_num = result["page_num"]
        
        if category not in category_groups:
            category_groups[category] = []
        
        if (category in always_separate_categories or 
            (category_groups[category] and 
             len(category_groups[category][-1]) >= get_max_pages_for_category(category))):
            category_groups[category].append([])
        
        if not category_groups[category]:
            category_groups[category].append([])
            
        category_groups[category][-1].append(result)
    
    for category, page_groups in category_groups.items():
        save_folder = os.path.join(output_path, category.replace(" ", "_"))
        os.makedirs(save_folder, exist_ok=True)
        
        for group_idx, page_group in enumerate(page_groups, 1):
            try:
                new_doc = fitz.open()
                
                page_nums = [p["page_num"] - 1 for p in page_group]
                
                for page_idx in page_nums:
                    new_doc.insert_pdf(doc, from_page=page_idx, to_page=page_idx)
                
                file_name = f"{category.replace(' ', '_')}_{session_id}_{group_idx}.pdf"
                file_path = os.path.join(save_folder, file_name)
                new_doc.save(file_path)
                new_doc.close()
                
                for page_result in page_group:
                    classification_log.append({
                        "Page Number": page_result["page_num"],
                        "Category": category,
                        "Quality Status": page_result["quality"],
                        "Processing Method": "Ultra-Fast Hybrid",
                        "Text Length": page_result["text_length"],
                        "Output File": file_name
                    })
            except Exception as e:
                logger.error(f"Error creating PDF for {category} group {group_idx}: {e}")
    
    return classification_log

# ---------------- Ultra-Fast Main Classification Function ----------------
def classify_document_ultra_fast(file, progress=gr.Progress()) -> str:
    """
    Ultra-optimized classification targeting <12 seconds for 32 pages.
    """
    try:
        start_time = time.time()
        max_workers = None  # Initialize max_workers to None
        
        session_id = str(uuid.uuid4())[:8]  # Unique short UUID
        session_output_path = os.path.join(OUTPUT_BASE_PATH, f"session_{session_id}")
        os.makedirs(session_output_path, exist_ok=True)
        
        logger.info(f"Configuration check:")
        logger.info(f"   Loaded {len(classification_keywords)} categories from JSON")
        for cat, keywords in list(classification_keywords.items())[:3]:
            sample = [kw['keyword'] for kw in keywords[:3]]
            logger.info(f"   {cat}: {sample}..." if len(keywords) > 3 else f"   {cat}: {sample}")
        
        if not file:
            raise ValueError("No file uploaded")
        
        file_ext = file.name.lower().split('.')[-1]
        if file_ext not in ['pdf', 'png', 'jpg', 'jpeg', 'tiff', 'tif', 'docx', 'xlsx']:
            raise ValueError(f"Unsupported file type: {file_ext}")
        
        # Add file size limit (e.g., 100MB)
        if os.path.getsize(file.name) > 100 * 1024 * 1024:
            raise ValueError("File too large (max 100MB)")
        
        if file_ext in ['png', 'jpg', 'jpeg', 'tiff', 'tif']:
            logger.info("Processing image with ultra-fast pipeline...")
            
            processor = PageProcessor()
            try:
                img = Image.open(file.name)
            except Exception as e:
                raise ValueError(f"Failed to open image: {e}")
            
            for config in processor.tesseract_configs:
                try:
                    text = pytesseract.image_to_string(img, config=config)
                    if text.strip():
                        break
                except Exception as e:
                    logger.warning(f"OCR failed for image with config {config}: {e}")
            
            category = classify_text_ultra_fast(text, page_num="Image")
            quality = "Good" if text.strip() else "Poor Quality"
            
            save_folder = os.path.join(session_output_path, category.replace(" ", "_"))
            os.makedirs(save_folder, exist_ok=True)
            output_path = os.path.join(save_folder, os.path.basename(file.name))
            shutil.copy2(file.name, output_path)
            
            classification_log = [{
                "Page Number": "Image",
                "Category": category,
                "Quality Status": quality,
                "Processing Method": "Ultra-Fast Tesseract",
                "Text Length": len(text)
            }]

        elif file_ext == 'pdf':
            try:
                doc = fitz.open(file.name)
            except Exception as e:
                raise ValueError(f"Failed to open PDF: {e}")
            
            total_pages = len(doc)
            logger.info(f"Processing {total_pages} pages with ULTRA-FAST parallel pipeline...")
            
            if total_pages > 500:  # Arbitrary limit for robustness
                raise ValueError("PDF too large (max 500 pages)")
            
            cpu_count = os.cpu_count() or 4
            max_workers = min(cpu_count, max(1, total_pages // 8 + 1))
            
            progress(0, desc="Starting text extraction...")
            
            extract_start = time.time()
            
            batch_size = max(4, total_pages // max_workers)
            page_batches = []
            for i in range(0, total_pages, batch_size):
                batch_end = min(i + batch_size, total_pages)
                batch_pages = [(doc[j], j + 1) for j in range(i, batch_end)]
                page_batches.append(batch_pages)
            
            all_results = []
            processed_batches = 0
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                future_to_batch = {
                    executor.submit(process_page_batch, batch, session_output_path): batch_idx 
                    for batch_idx, batch in enumerate(page_batches)
                }
                
                for future in as_completed(future_to_batch):
                    batch_idx = future_to_batch[future]
                    try:
                        batch_results = future.result()
                        all_results.extend(batch_results)
                        processed_batches += 1
                        progress(processed_batches / len(page_batches), desc=f"Processed batch {processed_batches}/{len(page_batches)}")
                        logger.info(f"Batch {batch_idx + 1}/{len(page_batches)} completed")
                    except Exception as e:
                        logger.error(f"Batch {batch_idx + 1} failed: {e}")
            
            all_results.sort(key=lambda x: x["page_num"])
            
            extract_time = time.time() - extract_start
            logger.info(f"Text extraction completed in {extract_time:.2f}s")
            
            progress(1, desc="Creating output PDFs...")
            pdf_start = time.time()
            classification_log = create_pdfs_batch_optimized(doc, all_results, session_output_path, session_id)
            pdf_time = time.time() - pdf_start
            logger.info(f"PDF creation completed in {pdf_time:.2f}s")
            
            doc.close()

        elif file_ext == 'docx':
            logger.info("Processing DOCX with ultra-fast pipeline...")
            try:
                doc = Document(file.name)
            except Exception as e:
                raise ValueError(f"Failed to open DOCX: {e}")
            text = "\n".join([para.text for para in doc.paragraphs])
            category = classify_text_ultra_fast(text, page_num="DOCX")
            quality = "Good" if text.strip() else "Poor Quality"
            
            save_folder = os.path.join(session_output_path, category.replace(" ", "_"))
            os.makedirs(save_folder, exist_ok=True)
            output_path = os.path.join(save_folder, os.path.basename(file.name))
            shutil.copy2(file.name, output_path)
            
            classification_log = [{
                "Page Number": "DOCX",
                "Category": category,
                "Quality Status": quality,
                "Processing Method": "Ultra-Fast DOCX",
                "Text Length": len(text)
            }]

        elif file_ext == 'xlsx':
            logger.info("Processing XLSX with ultra-fast pipeline...")
            try:
                wb = load_workbook(file.name)
            except Exception as e:
                raise ValueError(f"Failed to open XLSX: {e}")
            text = ""
            for sheet in wb:
                for row in sheet.iter_rows(values_only=True):
                    text += " ".join([str(cell) for cell in row if cell is not None]) + "\n"
            category = classify_text_ultra_fast(text, page_num="XLSX")
            quality = "Good" if text.strip() else "Poor Quality"
            
            save_folder = os.path.join(session_output_path, category.replace(" ", "_"))
            os.makedirs(save_folder, exist_ok=True)
            output_path = os.path.join(save_folder, os.path.basename(file.name))
            shutil.copy2(file.name, output_path)
            
            classification_log = [{
                "Page Number": "XLSX",
                "Category": category,
                "Quality Status": quality,
                "Processing Method": "Ultra-Fast XLSX",
                "Text Length": len(text)
            }]

        total_time = time.time() - start_time
        log_file_path = os.path.join(session_output_path, "classification_log.json")
        
        pages_processed = len(classification_log) if file_ext == 'pdf' else 1
        pages_per_second = round(pages_processed / total_time, 2) if total_time > 0 else 0
        
        log_data = {
            "processing_time_seconds": round(total_time, 2),
            "total_pages": pages_processed,
            "pages_per_second": pages_per_second,
            "performance_target": "< 12 seconds for 32 pages",
            "optimization_level": "Ultra-Fast Parallel",
            "session_id": session_id,
            "results": classification_log
        }
        
        with open(log_file_path, "w", encoding="utf-8") as f:
            json.dump(log_data, f, indent=4)

        target_time = (pages_processed / 32) * 12 if pages_processed > 0 else 12.0
        performance_status = "TARGET ACHIEVED!" if total_time < target_time else "Above target"
        speedup_vs_23s = round(23 / total_time, 1) if total_time > 0 else 0
        
        # Conditionally include parallel processing info only for PDFs
        optimizations = [
            "Zero-file I/O OCR (memory-based)",
            "Batch PDF operations" if file_ext == 'pdf' else "Direct file processing",
            "Pre-compiled keyword matching with weights",
            "Smart page type routing",
            "Enhanced preprocessing for low-quality scans",
            "DOCX/XLSX support"
        ]
        
        # Define newline join string for f-string
        newline_join = '\n   • '
        if file_ext == 'pdf' and max_workers is not None:
            optimizations.insert(0, f"Parallel processing ({max_workers} workers)")
        
        return f"""✅ ULTRA-FAST Classification Completed!

⏱️  **Performance Results:**
   • Total Time: {total_time:.1f} seconds
   • Speed: {pages_per_second:.1f} pages/second
   • {performance_status} (Target: <{target_time:.1f}s for {pages_processed} pages)
   • Speedup vs Original: {speedup_vs_23s}x faster
   • Session ID: {session_id}

🚀 **Optimizations Applied:**
   • {newline_join.join(optimizations)}

📁 **Output:** {session_output_path}
📊 **Log:** {log_file_path}"""

    except Exception as e:
        logger.error(f"Classification error: {str(e)}")
        raise gr.Error(f"Classification failed: {str(e)}")
    finally:
        gc.collect()

def should_start_new_pdf(category: str, prev_category: str, current_pages_count: int) -> bool:
    """
    Optimized PDF splitting logic based on config.
    """
    if category != prev_category:
        return True
    
    if category in always_separate_categories:
        return True
    
    max_pages = get_max_pages_for_category(category)
    if current_pages_count >= max_pages:
        return True
    
    return False

# ---------------- Gradio UI ----------------
def create_interface():
    """
    Create ultra-fast Gradio interface with progress tracking.
    """
    with gr.Blocks() as demo:
        gr.Markdown("Please upload the document. Document should be below 100mb.")

        with gr.Row():
            file_input = gr.File(
                label="Upload File (PDF / Image / TIFF / DOCX / XLSX)", 
                file_types=[".pdf", ".png", ".jpg", ".jpeg", ".tiff", ".tif", ".docx", ".xlsx"]
            )

        with gr.Row():
            classify_btn = gr.Button("⚡ ULTRA-FAST Classify", variant="primary", size="lg")
            
        with gr.Row():
            output_text = gr.Textbox(
                label="⚡ Ultra-Fast Results & Performance Analysis", 
                interactive=False, 
                lines=15,
                placeholder="Upload a document and click 'ULTRA-FAST Classify' to see results..."
            )

        classify_btn.click(
            fn=classify_document_ultra_fast, 
            inputs=file_input, 
            outputs=output_text
        )
    
    return demo

# ---------------- Launch Application ----------------
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Ultra-Fast Document Classification")
    parser.add_argument("--config", type=str, help="Path to configuration JSON")
    args = parser.parse_args()
    
    if args.config:
        CONFIG_PATH_ENV = args.config
    
    logger.info("Starting ULTRA-FAST Document Classification System...")
    logger.info("Target: <12 seconds for 32 pages (50%+ speed improvement)")
    logger.info("Parallel processing + Memory-based OCR + Batch operations + Fully Configurable + DOCX/XLSX Support")
    
    demo = create_interface()
    if demo:
        demo.launch()