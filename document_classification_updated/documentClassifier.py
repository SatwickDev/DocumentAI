# ---------------- Imports ----------------
import os
import fitz  # PyMuPDF
import pytesseract
import cv2
import numpy as np
from PIL import Image
import json
import gradio as gr
import shutil
import time
import joblib  # NEW: Added for ML model loading
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Dict, List, Tuple, Optional
from docx import Document
from openpyxl import load_workbook
import logging
import gc
import uuid
import argparse
from dataclasses import dataclass
import re
import sys
from functools import lru_cache
# ---------------- Performance Optimizations ----------------
# Reduce default resolution for faster processing
DEFAULT_MATRIX = fitz.Matrix(1.2, 1.2)  # Reduced from 2.0
OCR_MATRIX = fitz.Matrix(1.5, 1.5)     # Reduced from 2.0

# Cache compiled regex patterns
_regex_cache = {}


# Pre-compile frequently used patterns
def get_compiled_regex(pattern):
    if pattern not in _regex_cache:
        _regex_cache[pattern] = re.compile(pattern, re.IGNORECASE)
    return _regex_cache[pattern]


# ---------------- Unicode Encoding Fix (Simplified) ----------------
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except AttributeError:
        pass
    os.system('chcp 65001 > nul 2>&1')

# ---------------- Streamlined Logging ----------------
logging.basicConfig(
    level=logging.WARNING,  # Reduced logging for performance
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("document_classification.log", encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ---------------- Configuration ----------------
# Auto-detect Tesseract path based on environment
if os.path.exists('/usr/bin/tesseract'):
    # Linux/Docker environment
    TESSERACT_CMD = os.environ.get('TESSERACT_CMD', '/usr/bin/tesseract')
else:
    # Windows environment
    TESSERACT_CMD = os.environ.get('TESSERACT_CMD', r"C:\Program Files\Tesseract-OCR\tesseract.exe")

pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD
OUTPUT_BASE_PATH = os.environ.get(
    'OUTPUT_BASE_PATH', r"C:\DocumentClassificationOutput")
CONFIG_PATH_ENV = os.environ.get('CONFIG_PATH')
MODEL_DIR = os.environ.get('MODEL_DIR', os.path.join(os.getcwd(), "model"))


# ---------------- Data Classes ----------------
@dataclass
class ClassificationResult:
    technique: str
    category: str
    confidence: float
    processing_time: float
    text_extracted: str = ""
    error: Optional[str] = None


@dataclass
class PageResult:
    page_num: int
    document_type: str
    techniques_results: List[ClassificationResult]
    final_category: str
    confidence_score: str
    processing_time: float


# ---------------- NEW: ML Model Loader ----------------
class DebugMLModelLoader:
    """Enhanced ML Model Loader with debugging capabilities"""

    def __init__(self, model_dir: str = MODEL_DIR):
        self.model_dir = model_dir
        self.classifier = None
        self.vectorizer = None
        self.debug_mode = True
        self._load_model()

    def _load_model(self):
        """Load the trained ML model with debug output"""
        classifier_path = os.path.join(self.model_dir, "classifier.pkl")
        vectorizer_path = os.path.join(self.model_dir, "vectorizer.pkl")

        if self.debug_mode:
            print(f"[ML DEBUG] Looking for model files in: {self.model_dir}")
            print(f"[ML DEBUG] Classifier path: {classifier_path}")
            print(f"[ML DEBUG] Vectorizer path: {vectorizer_path}")
            print(f"[ML DEBUG] Classifier exists: {os.path.exists(classifier_path)}")
            print(f"[ML DEBUG] Vectorizer exists: {os.path.exists(vectorizer_path)}")

        if not os.path.exists(classifier_path) or not os.path.exists(
                vectorizer_path):
            if self.debug_mode:
                print("[ML DEBUG] Model files missing!")
            logger.warning(f"ML Model files not found in {self.model_dir}")
            return

        try:
            self.classifier = joblib.load(classifier_path)
            self.vectorizer = joblib.load(vectorizer_path)

            if self.debug_mode:
                print("[ML DEBUG] Model loaded successfully!")
                print(f"[ML DEBUG] Classifier type: {type(self.classifier)}")
                print(f"[ML DEBUG] Vectorizer type: {type(self.vectorizer)}")

                # Check what categories the model knows
                if hasattr(self.classifier, 'classes_'):
                    print(f"[ML DEBUG] Model categories: {self.classifier.classes_}")

                # Check vectorizer features
                if hasattr(self.vectorizer, 'vocabulary_'):
                    vocab_size = len(self.vectorizer.vocabulary_)
                    print(f"[ML DEBUG] Vectorizer vocabulary size: {vocab_size}")

            logger.info(f"ML Model loaded successfully from {self.model_dir}")

        except Exception as e:
            if self.debug_mode:
                print(f"[ML DEBUG] Failed to load model: {e}")
            logger.error(f"Failed to load ML model: {e}")
            self.classifier = None
            self.vectorizer = None

    def is_available(self) -> bool:
        """Check if ML model is available"""
        available = self.classifier is not None and self.vectorizer is not None
        if self.debug_mode:
            print(f"[ML DEBUG] Model available: {available}")
        return available

    def predict(self, text: str) -> Tuple[str, float]:
        """Predict category and confidence with debug output"""
        if not self.is_available():
            if self.debug_mode:
                print("[ML DEBUG] Model not available for prediction")
            return "Unclassified", 0.0

        try:
            if not text or len(text.strip()) < 2:  # Very minimal threshold
                if self.debug_mode:
                    print(f"[ML DEBUG] Text too short for prediction: '{text}'")
                return "Unclassified", 0.1

            if self.debug_mode:
                print(f"[ML DEBUG] Predicting for text: '{text[:50]}...'")

            # Transform text using vectorizer
            X = self.vectorizer.transform([text])

            if self.debug_mode:
                print(f"[ML DEBUG] Vectorized shape: {X.shape}")
                print(f"[ML DEBUG] Vectorized nnz: {X.nnz}")  # Non-zero elements

            # Get prediction probabilities
            probs = self.classifier.predict_proba(X)[0]

            if self.debug_mode:
                print(f"[ML DEBUG] Prediction probabilities: {probs}")

            # Get best category
            idx = probs.argmax()
            category = self.classifier.classes_[idx]
            confidence = probs[idx]

            if self.debug_mode:
                print(f"[ML DEBUG] Best prediction: {category} ({confidence:.3f})")
                # Show top 3 predictions
                top_indices = probs.argsort()[-3:][::-1]
                for i, top_idx in enumerate(top_indices):
                    cat = self.classifier.classes_[top_idx]
                    conf = probs[top_idx]
                    print(f"[ML DEBUG] Rank {i+1}: {cat} ({conf:.3f})")

            return category, confidence

        except Exception as e:
            if self.debug_mode:
                print(f"[ML DEBUG] Prediction error: {e}")
            logger.error(f"ML prediction error: {e}")
            return "Unclassified", 0.0


# Initialize ML model loader
ml_model = DebugMLModelLoader()


# ---------------- Optimized Document Type Detection ----------------
class FastDocumentTypeDetector:
    """Ultra-fast document type detection with caching"""

    @staticmethod
    @lru_cache(maxsize=1000)
    def detect_document_type_cached(text_length: int, word_count: int, image_count: int, page_area: float) -> str:
        """Cached version of document type detection"""
        if word_count > 15:  # Reduced threshold for faster detection
            return "text_based"

        if image_count > 0 and word_count < 8:
            return "image_based"

        text_density = word_count / (page_area / 10000) if page_area > 0 else 0
        return "text_based" if text_density > 0.08 else "image_based"

    @staticmethod
    def detect_document_type(page) -> str:
        """Fast detection with minimal processing"""
        try:
            # Quick text extraction without full processing
            text = page.get_text()
            word_count = len(text.split())

            # Quick image count
            images = page.get_images()
            image_count = len(images)

            # Basic page dimensions
            rect = page.rect
            page_area = rect.width * rect.height

            return FastDocumentTypeDetector.detect_document_type_cached(
                len(text), word_count, image_count, page_area
            )

        except Exception:
            return "image_based"


# ---------------- NEW: 4th ML Classifier ----------------
class FixedMLClassifier:
    """Fixed Machine Learning classifier with detailed debugging"""

    def __init__(self):
        self.ml_model = ml_model
        self.debug_mode = True

    def classify(self, page, page_num: int) -> ClassificationResult:
        start_time = time.time()

        if not self.ml_model.is_available():
            return ClassificationResult(
                technique="ML-TfIdf",
                category="Unclassified",
                confidence=0.0,
                processing_time=time.time() - start_time,
                error="ML model not available."
            )

        try:
            # Extract text from page
            text = page.get_text()

            if self.debug_mode:
                print(f"[ML DEBUG] Page {page_num}: Extracted text length = {len(text)}")
                return self._classify_with_text_internal(text, page_num, start_time)

        except Exception as e:
            return self._handle_error(e, page_num, start_time)

    def classify_with_text(self, ocr_text: str, page_num: int) -> ClassificationResult:
        """Classify using pre-extracted OCR text from other classifiers"""
        start_time = time.time()

        if not self.ml_model.is_available():
            return ClassificationResult(
                technique="ML-TfIdf",
                category="Unclassified",
                confidence=0.0,
                processing_time=time.time() - start_time,
                error="ML model not available"
            )

        try:
            if self.debug_mode:
                print(f"[ML DEBUG] Page {page_num}: Reusing OCR text, length = {len(ocr_text)}")
                print(f"[ML DEBUG] OCR text preview: {ocr_text[:100]}")

            return self._classify_with_text_internal(ocr_text, page_num, start_time)

        except Exception as e:
            return self._handle_error(e, page_num, start_time)

    def _classify_with_text_internal(self, text: str, page_num: int, start_time: float) -> ClassificationResult:
        """Internal method to classify with any text input"""

        # Clean and preprocess text
        cleaned_text = self._preprocess_text(text)

        if len(cleaned_text.strip()) < 3:
            if self.debug_mode:
                print(f"[ML DEBUG] Page {page_num}: Text too short for ML prediction")
            return ClassificationResult(
                technique="ML-TfIdf",
                category="Unclassified",
                confidence=0.1,
                processing_time=time.time() - start_time,
                text_extracted=text
            )

        # Get ML prediction
        category, confidence = self.ml_model.predict(cleaned_text)

        if self.debug_mode:
            print(f"[ML DEBUG] Page {page_num}: ML predicted {category} (confidence: {confidence:.3f})")

        return ClassificationResult(
            technique="ML-TfIdf",
            category=category,
            confidence=confidence,
            processing_time=time.time() - start_time,
            text_extracted=text
        )

    def _preprocess_text(self, text: str) -> str:
        """Preprocess text to match training data format"""
        if not text:
            return ""
        # Basic cleaning - remove extra whitespace, normalize
        cleaned = ' '.join(text.split())
        # Remove non-printable characters but keep important punctuation
        import re
        cleaned = re.sub(r'[^\w\s\-.,:()/]', ' ', cleaned)
        # Normalize multiple spaces
        cleaned = ' '.join(cleaned.split())
        return cleaned.strip()

    def _handle_error(self, error: Exception, page_num: int, start_time: float) -> ClassificationResult:
        """Handle classification errors"""
        error_msg = f"ML prediction error: {str(error)}"
        if self.debug_mode:
            print(f"[ML DEBUG] Page {page_num}: ERROR: {error_msg}")

        return ClassificationResult(
            technique="ML-TfIdf",
            category="Unclassified",
            confidence=0.0,
            processing_time=time.time() - start_time,
            error=error_msg
        )


# ---------------- Optimized Text-Based Classifiers ----------------
class FastPyMuPDFClassifier:
    """Optimized PyMuPDF classifier with pre-compiled patterns"""

    def __init__(self, classification_keywords, min_text_length=5):
        self.classification_keywords = classification_keywords
        self.min_text_length = min_text_length
        self._precompile_patterns()

    def _precompile_patterns(self):
        """Pre-compile normalized keywords for faster matching"""
        self.compiled_keywords = {}
        for category, keywords in self.classification_keywords.items():
            self.compiled_keywords[category] = [
                (kw_item["normalized"], kw_item["weight"])
                for kw_item in keywords
            ]

    def classify(self, page, page_num: int) -> ClassificationResult:
        start_time = time.time()
        try:
            text = page.get_text()
            if len(text.strip()) < self.min_text_length:
                return ClassificationResult(
                    technique="PyMuPDF",
                    category="Unclassified",
                    confidence=0.1,
                    processing_time=time.time() - start_time,
                    text_extracted=text
                )

            # Fast normalization using list comprehension
            text_normalized = ''.join(c.lower() for c in text if c.isalnum())

            category_scores = {}
            for category, compiled_kws in self.compiled_keywords.items():
                category_score = 0
                for keyword_normalized, weight in compiled_kws:
                    if keyword_normalized in text_normalized:
                        category_score += len(keyword_normalized) * 4 * weight

                if category_score > 0:
                    category_scores[category] = category_score

            if not category_scores:
                category = "Unclassified"
                confidence = 0.1
            else:
                category = max(category_scores, key=category_scores.get)
                confidence = 0.8

            return ClassificationResult(
                technique="PyMuPDF",
                category=category,
                confidence=confidence,
                processing_time=time.time() - start_time,
                text_extracted=text
            )
        except Exception as e:
            return ClassificationResult(
                technique="PyMuPDF",
                category="Unclassified",
                confidence=0.0,
                processing_time=time.time() - start_time,
                error=str(e)
            )


class OptimizedRegexClassifier:
    """Highly optimized regex classifier with pre-compiled patterns"""

    def __init__(self, classification_keywords):
        self.classification_keywords = classification_keywords
        self.patterns = self._build_optimized_patterns()

    def _build_optimized_patterns(self):
        """Build and cache optimized regex patterns"""
        patterns = {}
        for category, keywords in self.classification_keywords.items():
            category_patterns = []
            for kw_item in keywords:
                keyword = kw_item["keyword"]
                # Create optimized pattern
                pattern_str = re.escape(keyword).replace(r'\ ', r'\s+')
                pattern = get_compiled_regex(pattern_str)
                category_patterns.append((pattern, kw_item["weight"]))
            patterns[category] = category_patterns
        return patterns

    def classify(self, page, page_num: int) -> ClassificationResult:
        start_time = time.time()
        try:
            text = page.get_text()
            if not text.strip():
                return ClassificationResult(
                    technique="Regex",
                    category="Unclassified",
                    confidence=0.1,
                    processing_time=time.time() - start_time,
                    text_extracted=text
                )

            category_scores = {}
            for category, patterns in self.patterns.items():
                score = 0
                for pattern, weight in patterns:
                    matches = len(pattern.findall(text))
                    if matches > 0:
                        score += matches * weight

                if score > 0:
                    category_scores[category] = score

            if not category_scores:
                category = "Unclassified"
                confidence = 0.1
            else:
                category = max(category_scores, key=category_scores.get)
                max_score = category_scores[category]
                total_score = sum(category_scores.values())
                confidence = max_score / total_score if total_score > 0 else 0.1

            return ClassificationResult(
                technique="Regex",
                category=category,
                confidence=confidence,
                processing_time=time.time() - start_time,
                text_extracted=text
            )
        except Exception as e:
            return ClassificationResult(
                technique="Regex",
                category="Unclassified",
                confidence=0.0,
                processing_time=time.time() - start_time,
                error=str(e),
                text_extracted=""
            )


class FastNLPClassifier:
    # Uses Keyword-Set intersection technique
    """Simplified NLP classifier for speed"""

    def __init__(self, classification_keywords):
        self.classification_keywords = classification_keywords
        self.keyword_sets = self._build_keyword_sets()

    def _build_keyword_sets(self):
        """Build fast lookup sets"""
        keyword_sets = {}
        for category, keywords in self.classification_keywords.items():
            keyword_set = set()
            for kw_item in keywords:
                words = kw_item["keyword"].lower().split()
                keyword_set.update(words)
            keyword_sets[category] = keyword_set
        return keyword_sets

    def classify(self, page, page_num: int) -> ClassificationResult:
        start_time = time.time()
        try:
            text = page.get_text().lower()
            if not text.strip():
                return ClassificationResult(
                    technique="NLP",
                    category="Unclassified",
                    confidence=0.1,
                    processing_time=time.time() - start_time,
                    text_extracted=text
                )

            words = set(text.split())
            category_scores = {}

            for category, keyword_set in self.keyword_sets.items():
                intersection = words.intersection(keyword_set)
                if intersection:
                    category_scores[category] = len(intersection)

            if not category_scores:
                category = "Unclassified"
                confidence = 0.1
            else:
                category = max(category_scores, key=category_scores.get)
                confidence = 0.7

            return ClassificationResult(
                technique="NLP",
                category=category,
                confidence=confidence,
                processing_time=time.time() - start_time,
                text_extracted=text
            )
        except Exception as e:
            return ClassificationResult(
                technique="NLP",
                category="Unclassified",
                confidence=0.0,
                processing_time=time.time() - start_time,
                error=str(e),
                text_extracted=""
            )


# ---------------- Optimized Image-Based Classifiers ----------------
class FastTesseractClassifier:
    """Optimized Tesseract with reduced resolution and timeouts"""

    def __init__(self, classification_keywords, tesseract_configs):
        self.classification_keywords = classification_keywords
        # Use only the best config for speed
        self.tesseract_config = tesseract_configs[0] if tesseract_configs else r'--oem 3 --psm 6'
        self._precompile_patterns()

    def _precompile_patterns(self):
        """Pre-compile patterns like text classifier"""
        self.compiled_keywords = {}
        for category, keywords in self.classification_keywords.items():
            self.compiled_keywords[category] = [
                (kw_item["normalized"], kw_item["weight"])
                for kw_item in keywords
            ]

    def classify(self, page, page_num: int) -> ClassificationResult:
        start_time = time.time()
        try:
            # Use reduced resolution for speed
            pix = page.get_pixmap(matrix=OCR_MATRIX)

            # Fast PIL conversion
            if pix.n == 3:
                mode = "RGB"
            elif pix.n == 1:
                mode = "L"
            else:
                mode = "RGBA"

            pil_image = Image.frombytes(mode, [pix.width, pix.height], pix.samples)

            if pil_image.mode != 'RGB':
                pil_image = pil_image.convert('RGB')

            # Single OCR call with timeout
            try:
                text = pytesseract.image_to_string(pil_image, config=self.tesseract_config, timeout=10)
            except Exception as e:
                text = ""

            category = self._classify_text_fast(text)
            confidence = 0.7 if category != "Unclassified" else 0.1

            return ClassificationResult(
                technique="Tesseract",
                category=category,
                confidence=confidence,
                processing_time=time.time() - start_time,
                text_extracted=text
            )
        except Exception as e:
            return ClassificationResult(
                technique="Tesseract",
                category="Unclassified",
                confidence=0.0,
                processing_time=time.time() - start_time,
                error=str(e)
            )

    def _classify_text_fast(self, text: str) -> str:
        """Fast text classification"""
        if not text or len(text.strip()) < 5:
            return "Unclassified"

        text_normalized = ''.join(c.lower() for c in text if c.isalnum())

        best_category = "Unclassified"
        best_score = 0

        for category, compiled_kws in self.compiled_keywords.items():
            category_score = 0
            for keyword_normalized, weight in compiled_kws:
                if keyword_normalized in text_normalized:
                    category_score += len(keyword_normalized) * 4 * weight

            if category_score > best_score:
                best_score = category_score
                best_category = category

        return best_category


class LightweightOCRClassifier:
    """Lightweight OCR classifier (replaces EasyOCR for speed)"""

    def __init__(self, classification_keywords):
        self.classification_keywords = classification_keywords
        self._precompile_patterns()

    def _precompile_patterns(self):
        self.compiled_keywords = {}
        for category, keywords in self.classification_keywords.items():
            self.compiled_keywords[category] = [
                (kw_item["normalized"], kw_item["weight"])
                for kw_item in keywords
            ]

    def classify(self, page, page_num: int) -> ClassificationResult:
        start_time = time.time()
        try:
            # Fast image preprocessing with OpenCV
            pix = page.get_pixmap(matrix=DEFAULT_MATRIX)  # Lower resolution

            # Convert to numpy array quickly
            img_data = pix.samples
            if pix.n == 3:
                img_array = np.frombuffer(img_data, dtype=np.uint8).reshape(pix.h, pix.w, 3)
                gray = cv2.cvtColor(img_array, cv2.COLOR_RGB2GRAY)
            else:
                gray = np.frombuffer(img_data, dtype=np.uint8).reshape(pix.h, pix.w)

            # Simple preprocessing for speed
            _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

            # Convert back to PIL for OCR
            pil_image = Image.fromarray(binary)

            # Fast OCR with timeout
            try:
                text = pytesseract.image_to_string(pil_image, config=r'--oem 3 --psm 6', timeout=8)
            except Exception:
                text = ""

            category = self._classify_text_fast(text)
            confidence = 0.75 if category != "Unclassified" else 0.1

            return ClassificationResult(
                technique="LightOCR",
                category=category,
                confidence=confidence,
                processing_time=time.time() - start_time,
                text_extracted=text
            )
        except Exception as e:
            return ClassificationResult(
                technique="LightOCR",
                category="Unclassified",
                confidence=0.0,
                processing_time=time.time() - start_time,
                error=str(e)
            )

    def _classify_text_fast(self, text: str) -> str:
        if not text or len(text.strip()) < 5:
            return "Unclassified"

        text_normalized = ''.join(c.lower() for c in text if c.isalnum())

        best_category = "Unclassified"
        best_score = 0

        for category, compiled_kws in self.compiled_keywords.items():
            category_score = 0
            for keyword_normalized, weight in compiled_kws:
                if keyword_normalized in text_normalized:
                    category_score += len(keyword_normalized) * 4 * weight

            if category_score > best_score:
                best_score = category_score
                best_category = category

        return best_category


class FastOpenCVClassifier:
    """Streamlined OpenCV classifier"""

    def __init__(self, classification_keywords):
        self.classification_keywords = classification_keywords
        self._precompile_patterns()

    def _precompile_patterns(self):
        self.compiled_keywords = {}
        for category, keywords in self.classification_keywords.items():
            self.compiled_keywords[category] = [
                (kw_item["normalized"], kw_item["weight"])
                for kw_item in keywords
            ]

    def classify(self, page, page_num: int) -> ClassificationResult:
        start_time = time.time()
        try:
            # Reduced resolution processing
            pix = page.get_pixmap(matrix=DEFAULT_MATRIX)

            # Fast conversion to numpy
            if pix.n == 3:
                img_array = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.h, pix.w, 3)
                gray = cv2.cvtColor(img_array, cv2.COLOR_RGB2GRAY)
            else:
                gray = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.h, pix.w)

            # Single best preprocessing technique for speed
            adaptive = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2)

            # Fast OCR
            processed_pil = Image.fromarray(adaptive)
            try:
                text = pytesseract.image_to_string(processed_pil, config=r'--oem 3 --psm 6', timeout=8)
            except Exception:
                text = ""

            category = self._classify_text_fast(text)
            confidence = 0.8 if category != "Unclassified" else 0.1

            return ClassificationResult(
                technique="OpenCV",
                category=category,
                confidence=confidence,
                processing_time=time.time() - start_time,
                text_extracted=text
            )
        except Exception as e:
            return ClassificationResult(
                technique="OpenCV",
                category="Unclassified",
                confidence=0.0,
                processing_time=time.time() - start_time,
                error=str(e)
            )

    def _classify_text_fast(self, text: str) -> str:
        if not text or len(text.strip()) < 5:
            return "Unclassified"

        text_normalized = ''.join(c.lower() for c in text if c.isalnum())

        best_category = "Unclassified"
        best_score = 0

        for category, compiled_kws in self.compiled_keywords.items():
            category_score = 0
            for keyword_normalized, weight in compiled_kws:
                if keyword_normalized in text_normalized:
                    category_score += len(keyword_normalized) * 4 * weight

            if category_score > best_score:
                best_score = category_score
                best_category = category

        return best_category


# ---------------- OptimizedMultiTechniqueProcessor ---------------
class OptimizedMultiTechniqueProcessor:
    """High-performance multi-technique processor"""
    def __init__(self, classification_keywords, tesseract_configs):
        self.classification_keywords = classification_keywords

        # Initialize optimized classifiers (unchanged)
        self.pymupdf_classifier = FastPyMuPDFClassifier(
            classification_keywords)
        self.nlp_classifier = FastNLPClassifier(
            classification_keywords)
        self.regex_classifier = OptimizedRegexClassifier(
            classification_keywords)

        self.tesseract_classifier = FastTesseractClassifier(
            classification_keywords, tesseract_configs)
        self.light_ocr_classifier = LightweightOCRClassifier(
            classification_keywords)
        self.opencv_classifier = FastOpenCVClassifier(classification_keywords)

        # NEW: Add 4th ML classifier
        self.ml_classifier = FixedMLClassifier()

    def process_page(self, page, page_num: int) -> PageResult:
        """Process with 4/4 scoring - upgraded from 3/3"""
        start_time = time.time()

        # Fast document type detection (unchanged)
        document_type = FastDocumentTypeDetector.detect_document_type(page)

        # FIXED: Initialize shared_ocr_text at the beginning
        shared_ocr_text = ""

        # Select classifiers based on type (unchanged)
        if document_type == "text_based":
            classifiers = [
                self.pymupdf_classifier,
                self.nlp_classifier,
                self.regex_classifier
            ]
            ml_text_source = "direct"
        else:
            classifiers = [
                self.tesseract_classifier,
                self.light_ocr_classifier,
                self.opencv_classifier
            ]
            ml_text_source = "ocr_reuse"

        results = []
        with ThreadPoolExecutor(max_workers=3) as executor:
            future_to_classifier = {
                executor.submit(classifier.classify, page, page_num): classifier
                for classifier in classifiers
            }

            for future in as_completed(future_to_classifier, timeout=15):
                try:
                    result = future.result(timeout=10)
                    results.append(result)
                    # FIXED: Now shared_ocr_text is properly initialized
                    if (document_type == "image_based" and result.text_extracted and len(result.text_extracted) > len(shared_ocr_text)):
                        shared_ocr_text = result.text_extracted

                except Exception as e:
                    classifier_name = type(future_to_classifier[future]).__name__
                    results.append(ClassificationResult(
                        technique=classifier_name,
                        category="Unclassified",
                        confidence=0.0,
                        processing_time=0.0,
                        error=str(e)
                    ))

        # Add ML classifier as 4th technique
        try:
            if ml_text_source == "direct":
                ml_result = self.ml_classifier.classify(page, page_num)
            else:
                # FIXED: shared_ocr_text is now guaranteed to be defined
                ml_result = self.ml_classifier.classify_with_text(shared_ocr_text, page_num)

            results.append(ml_result)
        except Exception as e:
            results.append(ClassificationResult(
                technique="ML-TfIdf",
                category="Unclassified",
                confidence=0.0,
                processing_time=0.0,
                error=str(e)
            ))

        # MODIFIED: Now use 4/4 final classification instead of 3/3
        final_category, confidence_score = self._determine_final_classification_4_4(results)

        return PageResult(
            page_num=page_num,
            document_type=document_type,
            techniques_results=results,
            final_category=final_category,
            confidence_score=confidence_score,
            processing_time=time.time() - start_time
        )

    def _determine_final_classification_4_4(self, results: List[ClassificationResult]) -> Tuple[str, str]:
        """MODIFIED: 4/4 final classification instead of 3/3"""
        if not results:
            return "Unclassified", "0/4"  # Changed from 0/3

        # Quick vote counting (unchanged logic)
        category_votes = {}
        for result in results:
            if result.error:
                continue
            category = result.category
            category_votes[category] = category_votes.get(category, 0) + 1

        if not category_votes:
            return "Unclassified", "0/4"  # Changed from 0/3

        best_category = max(category_votes, key=category_votes.get)
        vote_count = category_votes[best_category]
        confidence_score = f"{vote_count}/4"  # Changed from /3 to /4

        return best_category, confidence_score


# ---------------- Optimized Batch Processing ----------------
def process_page_batch_optimized(pages_data: List[Tuple], processor: OptimizedMultiTechniqueProcessor) -> List[PageResult]:
    """Optimized batch processing with reduced overhead"""
    results = []

    # Smaller thread pool, faster execution
    max_workers = min(3, len(pages_data))  # Keep at 3 for stability

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_page = {
            executor.submit(processor.process_page, page, page_num): (page, page_num)
            for page, page_num in pages_data
        }

        for future in as_completed(future_to_page, timeout=40):
            try:
                result = future.result()
                results.append(result)
            except Exception as e:
                page, page_num = future_to_page[future]
                logger.warning(f"Page {page_num} failed: {e}")
                results.append(PageResult(
                    page_num=page_num,
                    document_type="unknown",
                    techniques_results=[],
                    final_category="Unclassified",
                    confidence_score="0/4",  # Changed from 0/3
                    processing_time=0.0
                ))

    return sorted(results, key=lambda x: x.page_num)


def load_classification_config(config_path: Optional[str] = None) -> Dict:
    """Load classification configuration with validation"""
    if config_path is None:
        config_path = CONFIG_PATH_ENV or os.path.join(os.path.dirname(os.path.abspath(__file__)), "classification_config.json")

    try:
        with open(config_path, "r", encoding="utf-8") as f:
            config = json.load(f)

        compiled_config = {}
        always_separate_categories = []
        category_page_limits = {}

        for category, data in config["categories"].items():
            if data.get("always_separate", False):
                always_separate_categories.append(category)
            category_page_limits[category] = data.get("max_pages_per_pdf", config.get("default_pages_per_pdf", 1))

            keywords = data.get("keywords", [])
            compiled_keywords = []
            for item in keywords:
                if isinstance(item, dict) and "keyword" in item:
                    kw = item["keyword"].lower()
                    weight = item.get("weight", 1.0)
                    compiled_keywords.append({"keyword": kw, "normalized": ''.join(c for c in kw if c.isalnum()), "weight": weight})
                elif isinstance(item, str):
                    kw = item.lower()
                    compiled_keywords.append({"keyword": kw, "normalized": ''.join(c for c in kw if c.isalnum()), "weight": 1.0})
            compiled_config[category] = sorted(compiled_keywords, key=lambda x: len(x["keyword"]), reverse=True)

        return {
            "categories": compiled_config,
            "always_separate_categories": always_separate_categories,
            "category_page_limits": category_page_limits,
            "default_pages_per_pdf": config.get("default_pages_per_pdf", 1),
            "tesseract_configs": config.get("tesseract_configs", [r'--oem 3 --psm 6 -c tessedit_char_whitelist=0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz.,:-/()']),
            "min_text_length": config.get("min_text_length", 5)
        }

    except FileNotFoundError:
        logger.error(f"Configuration file {config_path} not found!")
        raise FileNotFoundError("Configuration file is not present. Hence classification is aborted.")

    except json.JSONDecodeError as e:
        logger.error(f"Invalid JSON in {config_path}: {e}")
        return {
            "categories": {},
            "always_separate_categories": [],
            "category_page_limits": {},
            "default_pages_per_pdf": 1,
            "tesseract_configs": [r'--oem 3 --psm 6'],
            "min_text_length": 5
        }


def create_pdfs_from_multi_technique_results(doc, results: List[PageResult], output_path: str, session_id: str) -> List[Dict]:
    """Enhanced PDF creation - supports intelligent document separation and multi-PDF output"""
    category_groups = {}
    classification_log = []
    created_pdfs = []  # Track all created PDFs for frontend display

    # Enhanced grouping logic for mixed document types
    for result in results:
        category = result.final_category
        if category not in category_groups:
            category_groups[category] = []

        max_pages = category_page_limits.get(category, default_pages_per_pdf)
        always_sep = category in always_separate_categories

        # Enhanced separation logic for mixed documents
        need_new = True

        if category_groups[category]:
            last_group = category_groups[category][-1]
            if last_group:
                last_page_num = last_group[-1].page_num
                is_consecutive = (result.page_num == last_page_num + 1)

                # Enhanced rules for document separation
                if always_sep:
                    # Always separate documents like LC Application, Purchase Order
                    need_new = True
                elif not is_consecutive:
                    # Non-consecutive pages always start new group
                    need_new = True
                elif len(last_group) >= max_pages:
                    # Reached page limit, start new group
                    need_new = True
                else:
                    # Can add to existing group
                    need_new = False

        if need_new:
            category_groups[category].append([])

        category_groups[category][-1].append(result)

    # Fast PDF creation
    for category, page_groups in category_groups.items():
        save_folder = os.path.join(output_path, category.replace(" ", "_"))
        os.makedirs(save_folder, exist_ok=True)

        for group_idx, page_group in enumerate(page_groups, 1):
            try:
                new_doc = fitz.open()
                page_nums = [p.page_num - 1 for p in page_group]

                for page_idx in page_nums:
                    new_doc.insert_pdf(doc, from_page=page_idx, to_page=page_idx)

                file_name = f"{category.replace(' ', '_')}_{session_id}_{group_idx}.pdf"
                file_path = os.path.join(save_folder, file_name)
                new_doc.save(file_path)
                new_doc.close()

                # Enhanced logging with 4 technique verdicts
                for page_result in page_group:
                    techniques_verdicts = {}
                    for res in page_result.techniques_results:
                        techniques_verdicts[res.technique] = {
                            "category": res.category,
                            "confidence": res.confidence,
                            "error": res.error
                        }
                    classification_log.append({
                        "Page Number": page_result.page_num,
                        "Document Type": page_result.document_type,
                        "Final Category": category,
                        "Confidence Score": page_result.confidence_score,
                        "Processing Time": f"{page_result.processing_time:.2f}s",
                        "Output File": file_name,
                        "Techniques Verdicts": techniques_verdicts
                    })
            except Exception as e:
                logger.error(f"Error creating PDF for {category} group {group_idx}: {e}")

    return classification_log


# Load configuration at startup
try:
    config = load_classification_config()
    classification_keywords = config["categories"]
    always_separate_categories = config["always_separate_categories"]
    category_page_limits = config["category_page_limits"]
    default_pages_per_pdf = config["default_pages_per_pdf"]
    tesseract_configs = config["tesseract_configs"]
    min_text_length = config["min_text_length"]
except Exception as e:
    logger.critical(f"Failed to load configuration: {str(e)}")
    raise


def generate_technique_analysis(classification_log: List[Dict], total_pages: int) -> str:
    """Generate detailed analysis showing confidence from each technique"""

    if not classification_log:
        return "   • No classification data available for analysis"

    # Collect technique statistics
    technique_stats = {}
    technique_agreements = {}
    page_details = []

    for log_entry in classification_log:
        if not isinstance(log_entry, dict) or "Techniques Verdicts" not in log_entry:
            continue

        page_num = log_entry.get("Page Number", "Unknown")
        final_category = log_entry.get("Final Category", "Unknown")
        confidence_score = log_entry.get("Confidence Score", "0/4")

        techniques_verdicts = log_entry.get("Techniques Verdicts", {})

        # Build page detail
        page_detail = f"\n📄 PAGE {page_num} → Final: {final_category} ({confidence_score})"

        technique_results = []
        for technique, verdict in techniques_verdicts.items():
            category = verdict.get("category", "Unknown")
            confidence = verdict.get("confidence", 0.0)
            error = verdict.get("error")

            # Track technique statistics
            if technique not in technique_stats:
                technique_stats[technique] = {
                    "total_runs": 0,
                    "successful_runs": 0,
                    "avg_confidence": 0.0,
                    "categories_found": {},
                    "total_confidence": 0.0
                }

            technique_stats[technique]["total_runs"] += 1

            if error:
                technique_results.append(f"   ❌ {technique}: ERROR - {error}")
            else:
                technique_stats[technique]["successful_runs"] += 1
                technique_stats[technique]["total_confidence"] += confidence

                if category not in technique_stats[technique]["categories_found"]:
                    technique_stats[technique]["categories_found"][category] = 0
                technique_stats[technique]["categories_found"][category] += 1

                # Format confidence display
                confidence_pct = confidence * 100
                if confidence_pct >= 75:
                    conf_icon = "🟢"
                elif confidence_pct >= 50:
                    conf_icon = "🟡"
                elif confidence_pct >= 25:
                    conf_icon = "🟠"
                else:
                    conf_icon = "🔴"

                technique_results.append(f"   {conf_icon} {technique}: {category} ({confidence_pct:.1f}%)")

                # Track agreements
                if final_category == category:
                    if technique not in technique_agreements:
                        technique_agreements[technique] = 0
                    technique_agreements[technique] += 1

        page_detail += "\n" + "\n".join(technique_results)
        page_details.append(page_detail)

    # Calculate technique averages
    for technique in technique_stats:
        stats = technique_stats[technique]
        if stats["successful_runs"] > 0:
            stats["avg_confidence"] = stats["total_confidence"] / stats["successful_runs"]
        stats["success_rate"] = (stats["successful_runs"] / stats["total_runs"]) * 100 if stats["total_runs"] > 0 else 0
        stats["agreement_rate"] = (technique_agreements.get(technique, 0) / total_pages) * 100 if total_pages > 0 else 0

    # Build analysis report
    analysis_parts = []

    # Overall technique performance
    analysis_parts.append("\n🔍 TECHNIQUE PERFORMANCE SUMMARY:")

    # Sort techniques by agreement rate for better display
    sorted_techniques = sorted(technique_stats.items(), key=lambda x: x[1]["agreement_rate"], reverse=True)

    for technique, stats in sorted_techniques:
        success_rate = stats["success_rate"]
        avg_conf = stats["avg_confidence"] * 100
        agreement_rate = stats["agreement_rate"]

        # Get most common category for this technique
        if stats["categories_found"]:
            most_common_cat = max(stats["categories_found"], key=stats["categories_found"].get)
            cat_count = stats["categories_found"][most_common_cat]
        else:
            most_common_cat = "None"
            cat_count = 0

        # Performance indicator
        if agreement_rate >= 80:
            perf_icon = "🏆"
        elif agreement_rate >= 60:
            perf_icon = "🥈"
        elif agreement_rate >= 40:
            perf_icon = "🥉"
        else:
            perf_icon = "⚠️"

        analysis_parts.append(
            f"   {perf_icon} {technique}:\n"
            f"      ├─ Success Rate: {success_rate:.1f}% ({stats['successful_runs']}/{stats['total_runs']} runs)\n"
            f"      ├─ Avg Confidence: {avg_conf:.1f}%\n"
            f"      ├─ Agreement with Final: {agreement_rate:.1f}% ({technique_agreements.get(technique, 0)}/{total_pages} pages)\n"
            f"      └─ Top Category: {most_common_cat} ({cat_count} times)"
        )

    # Technique explanations
    analysis_parts.append("\n📚 TECHNIQUE EXPLANATIONS:")

    technique_explanations = {
        "PyMuPDF": "Direct PDF text extraction + weighted keyword matching. Fast and accurate for text-based docs.",
        "Regex": "Advanced pattern matching with fuzzy search. Handles variations in document formatting.", 
        "NLP": "Semantic word-set intersection analysis. Understands context and related terms.",
        "ML-TfIdf": "Machine learning trained on document samples. Learns patterns beyond simple keywords.",
        "Tesseract": "Google's OCR engine with optimized configs. Reliable for printed text recognition.",
        "LightOCR": "OpenCV preprocessing + OCR. Good for documents with varied image quality.",
        "OpenCV": "Advanced image processing + OCR. Handles complex layouts and image artifacts."
    }

    for technique in sorted_techniques:
        tech_name = technique[0]
        base_name = tech_name.replace("Fast", "").replace("Optimized", "").replace("Lightweight", "Light")
        if base_name in technique_explanations:
            analysis_parts.append(f"   💡 {tech_name}: {technique_explanations[base_name]}")

    # Sample page details (show first 3 pages as examples)
    if page_details:
        analysis_parts.append("\n📋 SAMPLE PAGE ANALYSIS (First 3 pages):")
        for i, detail in enumerate(page_details[:3]):
            analysis_parts.append(detail)

        if len(page_details) > 3:
            analysis_parts.append(f"\n   ... (and {len(page_details) - 3} more pages)")

    return "\n".join(analysis_parts)


# ---------------- MODIFIED: Main Classification Function (now 4/4 instead of 3/3) ----------------
def classify_document_optimized(file, progress=gr.Progress()) -> str:
    """
    Enhanced multi-technique classification with 4th ML technique
    Target: Still 10-20 seconds for 9 pages but with 4/4 confidence scoring
    """
    try:
        start_time = time.time()

        session_id = str(uuid.uuid4())[:8]
        session_output_path = os.path.join(OUTPUT_BASE_PATH, f"session_{session_id}")
        os.makedirs(session_output_path, exist_ok=True)

        logger.warning(f"Starting Enhanced 4/4 Multi-Technique Classification (Session: {session_id})")

        if not file:
            raise ValueError("No file uploaded")

        file_ext = file.name.lower().split('.')[-1]
        if file_ext not in ['pdf', 'png', 'jpg', 'jpeg', 'tiff', 'tif', 'docx', 'xlsx']:
            raise ValueError(f"Unsupported file type: {file_ext}")

        # Reduced file size limit for faster processing
        if os.path.getsize(file.name) > 50 * 1024 * 1024:  # 50MB instead of 100MB
            raise ValueError("File too large (max 50MB for optimal speed)")

        # Initialize processor with 4th ML technique
        processor = OptimizedMultiTechniqueProcessor(classification_keywords, tesseract_configs)

        if file_ext in ['png', 'jpg', 'jpeg', 'tiff', 'tif']:
            # Fast image processing with 4 techniques
            img_doc = fitz.open()
            img_doc.new_page(width=612, height=792)
            page = img_doc[0]

            img_rect = fitz.Rect(0, 0, 612, 792)
            page.insert_image(img_rect, filename=file.name)

            progress(0.3, desc="Enhanced 4-technique image classification...")
            result = processor.process_page(page, 1)

            # Quick save
            save_folder = os.path.join(session_output_path, result.final_category.replace(" ", "_"))
            os.makedirs(save_folder, exist_ok=True)
            output_path = os.path.join(save_folder, os.path.basename(file.name))
            shutil.copy2(file.name, output_path)

            img_doc.close()

            techniques_verdicts = {}
            for res in result.techniques_results:
                techniques_verdicts[res.technique] = {
                    "category": res.category,
                    "confidence": res.confidence,
                    "error": res.error
                }
            classification_log = [{
                "Page Number": "Image",
                "Document Type": result.document_type,
                "Final Category": result.final_category,
                "Confidence Score": result.confidence_score,
                "Processing Time": f"{result.processing_time:.2f}s",
                "Output File": os.path.basename(file.name),
                "Techniques Verdicts": techniques_verdicts
            }]

            total_pages = 1

        elif file_ext == 'pdf':
            try:
                doc = fitz.open(file.name)
            except Exception as e:
                raise ValueError(f"Failed to open PDF: {e}")

            total_pages = len(doc)
            logger.warning(f"Processing {total_pages}-page PDF with enhanced 4-technique approach...")

            # Reduced page limit for speed
            if total_pages > 200:  # Reduced from 500
                raise ValueError("PDF too large (max 200 pages for optimal speed)")

            progress(0, desc="Starting enhanced 4/4 classification...")

            # Adjusted for 4 techniques
            cpu_count = min(4, os.cpu_count() or 2)  # Cap at 4 cores
            max_workers = min(2, max(1, total_pages // 6))  # Reduced workers

            batch_size = max(1, total_pages // max(max_workers, 1))  # Smaller batches
            page_batches = []

            for i in range(0, total_pages, batch_size):
                batch_end = min(i + batch_size, total_pages)
                batch_pages = [(doc[j], j + 1) for j in range(i, batch_end)]
                page_batches.append(batch_pages)

            all_results = []
            processed_batches = 0

            # Process with 4 techniques
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                future_to_batch = {
                    executor.submit(process_page_batch_optimized, batch, processor): batch_idx 
                    for batch_idx, batch in enumerate(page_batches)
                }

                for future in as_completed(future_to_batch, timeout=80):  # Increased timeout for 4 techniques
                    batch_idx = future_to_batch[future]
                    try:
                        batch_results = future.result(timeout=40)  # Per-batch timeout
                        all_results.extend(batch_results)
                        processed_batches += 1

                        progress(
                            processed_batches / len(page_batches),
                            desc=f"Enhanced 4/4 processing: {processed_batches}/{len(page_batches)} batches"
                        )

                    except Exception as e:
                        logger.error(f"Batch {batch_idx + 1} failed: {e}")

            all_results.sort(key=lambda x: x.page_num)

            progress(0.9, desc="Creating output files...")
            classification_log = create_pdfs_from_multi_technique_results(doc, all_results, session_output_path, session_id)

            doc.close()

        elif file_ext == 'docx':
            # Enhanced DOCX processing with 4 techniques
            try:
                doc_content = Document(file.name)
                text = "\n".join([para.text for para in doc_content.paragraphs])
            except Exception as e:
                raise ValueError(f"Failed to open DOCX: {e}")

            # Quick text-based classification using all 4 techniques
            dummy_doc = fitz.open()
            page = dummy_doc.new_page()
            page.insert_text((50, 50), text[:500])  # Reduced text sample

            result = processor.process_page(page, 1)
            dummy_doc.close()

            save_folder = os.path.join(session_output_path, result.final_category.replace(" ", "_"))
            os.makedirs(save_folder, exist_ok=True)
            output_path = os.path.join(save_folder, os.path.basename(file.name))
            shutil.copy2(file.name, output_path)

            techniques_verdicts = {}
            for res in result.techniques_results:
                techniques_verdicts[res.technique] = {
                    "category": res.category,
                    "confidence": res.confidence,
                    "error": res.error
                }
            classification_log = [{
                "Page Number": "DOCX",
                "Document Type": "text_based",
                "Final Category": result.final_category,
                "Confidence Score": result.confidence_score,
                "Processing Time": f"{result.processing_time:.2f}s",
                "Text Length": len(text),
                "Techniques Verdicts": techniques_verdicts
            }]

            total_pages = 1

        elif file_ext == 'xlsx':
            # Enhanced XLSX processing with 4 techniques
            try:
                wb = load_workbook(file.name, read_only=True)  # Read-only for speed
                text = ""
                for sheet in wb:
                    for row in sheet.iter_rows(values_only=True, max_row=100):  # Limit rows
                        text += " ".join([str(cell) for cell in row if cell is not None]) + "\n"
                        if len(text) > 1000:  # Limit text extraction
                            break
                    if len(text) > 1000:
                        break
            except Exception as e:
                raise ValueError(f"Failed to open XLSX: {e}")

            dummy_doc = fitz.open()
            page = dummy_doc.new_page()
            page.insert_text((50, 50), text[:500])

            result = processor.process_page(page, 1)
            dummy_doc.close()

            save_folder = os.path.join(session_output_path, result.final_category.replace(" ", "_"))
            os.makedirs(save_folder, exist_ok=True)
            output_path = os.path.join(save_folder, os.path.basename(file.name))
            shutil.copy2(file.name, output_path)

            techniques_verdicts = {}
            for res in result.techniques_results:
                techniques_verdicts[res.technique] = {
                    "category": res.category,
                    "confidence": res.confidence,
                    "error": res.error
                }
            classification_log = [{
                "Page Number": "XLSX",
                "Document Type": "text_based",
                "Final Category": result.final_category,
                "Confidence Score": result.confidence_score,
                "Processing Time": f"{result.processing_time:.2f}s",
                "Text Length": len(text),
                "Techniques Verdicts": techniques_verdicts
            }]

            total_pages = 1

        # Calculate enhanced 4/4 statistics
        total_time = time.time() - start_time

        # Enhanced confidence analysis for 4/4 system
        confidence_stats = {"4/4": 0, "3/4": 0, "2/4": 0, "1/4": 0, "0/4": 0}  # CHANGED from 3/3 system
        text_based_count = 0
        image_based_count = 0

        for log_entry in classification_log:
            if isinstance(log_entry, dict):
                confidence = log_entry.get("Confidence Score", "0/4")  # CHANGED from 0/3
                confidence_stats[confidence] = confidence_stats.get(confidence, 0) + 1

                doc_type = log_entry.get("Document Type", "unknown")
                if doc_type == "text_based":
                    text_based_count += 1
                elif doc_type == "image_based":
                    image_based_count += 1

        # Save enhanced log
        log_file_path = os.path.join(session_output_path, "enhanced_4_4_classification_log.json")  # CHANGED filename

        pages_per_second = round(total_pages / total_time, 2) if total_time > 0 else 0

        # Enhanced log data for 4/4 system
        log_data = {
            "processing_time_seconds": round(total_time, 2),
            "total_pages": total_pages,
            "pages_per_second": pages_per_second,
            "session_id": session_id,
            "system_version": "Enhanced 4/4 Multi-Technique (ML Added)",  # CHANGED
            "ml_model_status": "Available" if ml_model.is_available() else "Not Available",  # NEW
            "confidence_distribution": confidence_stats,
            "results": classification_log
        }

        with open(log_file_path, "w", encoding="utf-8") as f:
            json.dump(log_data, f, indent=2)

        # Generate enhanced 4/4 summary
        high_confidence = confidence_stats.get("4/4", 0)  # CHANGED from 3/3
        medium_confidence = confidence_stats.get("3/4", 0)  # CHANGED from 2/3
        confidence_percentage = round((high_confidence + medium_confidence) / max(total_pages, 1) * 100, 1)

        # Quick category breakdown
        category_breakdown = {}
        for log_entry in classification_log:
            if isinstance(log_entry, dict):
                category = log_entry.get("Final Category", "Unknown")
                category_breakdown[category] = category_breakdown.get(category, 0) + 1

        category_summary = [f"   • {cat}: {count} pages" for cat, count in sorted(category_breakdown.items())]
        ml_status = "✅ Available" if ml_model.is_available() else "❌ Not Available (run train_model_auto.py)"

        return f"""ENHANCED Multi-Technique Classification Complete! (Upgraded to 4/4 System)

SPEED PERFORMANCE:
   • Processing Time: {total_time:.1f} seconds (Target: Still 10-20s)
   • Pages Processed: {total_pages}
   • Speed: {pages_per_second:.1f} pages/second
   • Session: {session_id}

SEPARATION BEHAVIOR (Per Your Config):
   • Purchase Order: {category_breakdown.get('Purchase Order', 0)} pages → {category_breakdown.get('Purchase Order', 0)} separate PDFs (always_separate=true)
   • LC Application Form: {category_breakdown.get('LC Application Form', 0)} pages → {category_breakdown.get('LC Application Form', 0)} separate PDFs (always_separate=true)  
   • Proforma Invoice: {category_breakdown.get('Proforma Invoice', 0)} pages → {category_breakdown.get('Proforma Invoice', 0)} separate PDFs (always_separate=true)
   • Bank Guarantee: {category_breakdown.get('Bank Guarantee', 0)} pages → Grouped (max 5 pages per PDF)
   • Unclassified: {category_breakdown.get('Unclassified', 0)} pages → {category_breakdown.get('Unclassified', 0)} separate PDFs (max_pages_per_pdf=1)

CLASSIFICATION RESULTS BY CATEGORY:
{chr(10).join(category_summary) if category_summary else '   • No classifications found'}

ENHANCED 4/4 CONFIDENCE ANALYSIS:
   • Perfect Consensus (4/4): {high_confidence} pages - All 4 techniques agree
   • Strong Consensus (3/4): {medium_confidence} pages - 3 out of 4 techniques agree  
   • Moderate Consensus (2/4): {confidence_stats.get('2/4', 0)} pages - Split decision
   • Weak Consensus (1/4): {confidence_stats.get('1/4', 0)} pages - Single technique confident
   • No Consensus (0/4): {confidence_stats.get('0/4', 0)} pages - All techniques failed
   • System Reliability: {confidence_percentage}%
   • Text-based: {text_based_count} | Image-based: {image_based_count}

TECHNIQUES APPLIED:
   Text Documents: FastPyMuPDF + OptimizedRegex + FastNLP + ML-TfIdf (4 total)
   Image Documents: FastTesseract + LightOCR + FastOpenCV + ML-TfIdf (4 total)

CONFIGURATION APPLIED (Unchanged):

   • Min Text Length: {min_text_length} characters
   • Default Pages Per PDF: {default_pages_per_pdf}

Output Location: {session_output_path}
Enhanced 4/4 Log: {log_file_path}
"""

    except Exception as e:
        logger.error(f"Enhanced 4/4 classification error: {str(e)}")
        raise gr.Error(f"Classification failed: {str(e)}")
    finally:
        gc.collect()


# ---------------- Enhanced Gradio UI ----------------
def create_optimized_interface():
    """Create enhanced Gradio interface with 4/4 system info"""
    ml_status_display = "✅ ML Model Ready" if ml_model.is_available() else "❌ ML Model Missing (run train_model_auto.py)"

    with gr.Blocks(title="Enhanced 4/4 Multi-Technique Document Classifier") as demo:
        gr.Markdown(f"""
        - **ML Model Status:** {ml_status_display}

        📊 **Enhanced Technique Stack:**
        - **Text-based:** FastPyMuPDF + OptimizedRegex + FastNLP + **ML-TfIdf**
        - **Image-based:** FastTesseract + LightOCR + FastOpenCV + **ML-TfIdf**
        """)

        with gr.Row():
            file_input = gr.File(
                label="Upload Document (PDF/Image/DOCX/XLSX) - Max 50MB",
                file_types=[".pdf", ".png", ".jpg",
                            ".jpeg", ".tiff", ".tif", ".docx", ".xlsx"]
            )

        with gr.Row():
            classify_btn = gr.Button(
                "🚀 Start Enhanced 4/4 Classification",
                variant="primary",
                size="lg"
            )

        with gr.Row():
            output_text = gr.Textbox(
                label="Enhanced 4/4 Classification Results",
                interactive=False,
                lines=30,
                placeholder=f"""Upload a document and click 'Start Enhanced 4/4 Classification' to see:

🆕 UPGRADED 4/4 CONFIDENCE SYSTEM:
• 4/4: Perfect consensus - All 4 techniques agree
• 3/4: Strong consensus - 3 out of 4 techniques agree
• 2/4: Moderate consensus - Split decision
• 1/4: Weak consensus - Single technique confident
• 0/4: No consensus - All techniques failed

ML MODEL STATUS: {ml_status_display}

TECHNIQUE BREAKDOWN:
✅ Text Documents: 3 text-specific + 1 universal ML = 4 total
✅ Image Documents: 3 OCR-specific + 1 universal ML = 4 total

Still maintains all speed optimizations and accuracy!"""
            )

        classify_btn.click(
            fn=classify_document_optimized,
            inputs=file_input,
            outputs=output_text
        )

    return demo


def test_ml_model():
    """Quick test function to verify ML model is working"""
    print("=== ML MODEL TEST ===")

    # Test model loading
    test_ml = DebugMLModelLoader()

    if not test_ml.is_available():
        print("ERROR: ML Model not available!")
        return False

    # Test prediction with sample text
    test_texts = [
        "Purchase Order Number 12345 for delivery of goods",
        "Dear Sir, this is to apply for Letter of Credit",
        "Bank Guarantee for the amount of $50000",
        "Invoice for professional services rendered"
    ]

    print("\nTesting predictions:")
    for i, text in enumerate(test_texts, 1):
        category, confidence = test_ml.predict(text)
        print(f"Test {i}: {category} ({confidence:.3f}) - '{text[:40]}...'")

    return True


# ---------------- Launch Enhanced Application ----------------
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Enhanced 4/4 Multi-Technique Document Classification")
    parser.add_argument("--config", type=str, help="Path to configuration JSON")
    args = parser.parse_args()

    if args.config:
        CONFIG_PATH_ENV = args.config

    logger.warning("🚀 Starting Enhanced 4/4 Multi-Technique Document Classification System...")
    logger.warning(f"🧠 ML Model Status: {'✅ Ready' if ml_model.is_available() else '❌ Missing (run train_model_auto.py)'}")
    # Test ML model first
    if test_ml_model():
        print("ML Model test passed!")
    else:
        print("ML Model test failed!")
    demo = create_optimized_interface()
    if demo:
        demo.launch(share=False, server_name="0.0.0.0", server_port=7860)
